import csv
from django.http import HttpResponse
from django.contrib import admin
from openpyxl import Workbook
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from .models import Cooperado,Brand, Category, Product, Branch, Controle,Phone,Perifericos,Prestador
from datetime import date

#Funcionarios
class ProductResource(resources.ModelResource):
    is_inactive = fields.Field(attribute='is_inactive', default=False)
    class Meta:
        model = Cooperado

#Maquina, Celular etc
@admin.register(Cooperado)
#Quando precisar fazer a exportação ou importa usar o modo abaixo
class CooperadoAdmin(ImportExportModelAdmin):#ImportExportModelAdmin serve para usar o import/export dentro Admin
#@admin.register(Cooperado)
#class CooperadoAdmin(admin.ModelAdmin):
    list_display = ('name','mat','cpf','rg','is_active','is_inactive',)
    search_fields = ('name' ,)
    
#Prestador
@admin.register(Prestador)
class PrestadorAdmin(admin.ModelAdmin):
    list_display = ('title','is_active','is_inactive',)
    search_fields = ('title' ,)

#Marca
@admin.register(Brand)
class BrandAdmin(admin.ModelAdmin):
    list_display = ('name', 'description', 'created_at', 'updated_at')
    search_fields = ('name',)
    # list_filter = ('is_active',)

    def export_to_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="brands.csv"'
        writer = csv.writer(response)
        writer.writerow(['nome', 'ativo', 'descrição', 'cirado em', 'atualizado em'])
        for brand in queryset:
            writer.writerow([brand.name, brand.is_active, brand.description,
                             brand.created_at, brand.updated_at])
        return response

    export_to_csv.short_description = 'Exportar para CSV'
    actions = [export_to_csv]

#Departamento
@admin.register(Category)
class DepartamentoAdmin(admin.ModelAdmin):
    list_display = ('name', 'is_active', 'description', 'created_at', 'updated_at')
    search_fields = ('name',)
    list_filter = ('is_active',)

    def export_to_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="categories.csv"'
        writer = csv.writer(response)
        writer.writerow(['nome', 'ativo', 'descrição', 'cirado em', 'atualizado em'])
        for category in queryset:
            writer.writerow([category.name, category.is_active, category.description,
                             category.created_at, category.updated_at])
        return response

    export_to_csv.short_description = 'Exportar para CSV'
    actions = [export_to_csv]

#Maquina, Celular etc
class ProductResource(resources.ModelResource):
    class Meta:
        model = Product

@admin.register(Product)
class ProductAdmin(ImportExportModelAdmin):#ImportExportModelAdmin serve para usar o import/export dentro Admin
#class ProductAdmin(admin.ModelAdmin):
    # resource_classes = [ProductResource]
    list_display = ('title', 'brand','category','processor','memory_ram','storage','description',)
    search_fields = ('title', 'brand__name', 'category__name',)
    list_filter = ('is_active', 'brand', 'category')
    
    #importando para excel
    def export_product_to_excel(request,self,queryset):
    # Cria o workbook e a planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Computadores"

        # Adiciona o cabeçalho
        headers = ['title', 'brand','category','processor','memory_ram','storage','description',]
        worksheet.append(headers)

        # Recupera os dados do modelo e preenche a planilha
        computer = Product.objects.all() #Busca em Controle todos o objetos
        for computers in computer: #Percorre os objetos
            worksheet.append([
                str (computers.title),
                str (computers.brand),
                str (computers.patrimonio),
                str (computers.category),
                str (computers.processor),
                str (computers.memory_ram),
                str (computers.storage),
                str (computers.description),
            #   controle.delivery.strftime("%Y-%m-%d"),
            #   controle.delivery.strftime("%H:%M:%S"),
            ])

        # Configura a resposta HTTP para o download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Perifericos.xlsx"
        workbook.save(response)
        return response
    
    export_product_to_excel.short_description = 'Exportar para excel'
    actions = [export_product_to_excel]

    # def export_to_csv(self, request, queryset):
    #     response = HttpResponse(content_type='text/csv')
    #     response['Content-Disposition'] = 'attachment; filename="products.csv"'
    #     writer = csv.writer(response)
    #     writer.writerow(['título', 'marca', 'categoria', 'preço',
    #                      'ativo', 'descrição', 'criado em', 'atualizado em'])
    #     for product in queryset:
    #         writer.writerow([product.title, product.brand.name, product.category.name,
    #                          product.price, product.is_active, product.description,
    #                          product.created_at, product.updated_at])
    #     return response

    # export_to_csv.short_description = 'Exportar para CSV'
    # actions = [export_to_csv]

# Filiais   
class BranchResource(resources.ModelResource):
    class Meta:
        model = Branch

@admin.register(Branch)
class BranchAdmin(ImportExportModelAdmin):
    resource_classes = [BranchResource]
    list_display = ('name','created_at','updated_at',)
    search_fields = ('name',)

#Maquina, Celular etc
class ProductResource(resources.ModelResource):
    class Meta:
        model = Controle

#Controles de notebooks e celular
@admin.register(Controle)
class ControleAdmin(ImportExportModelAdmin):#ImportExportModelAdmin serve para usar o import/export dentro Admin

    list_display = ('name','laptop','phones','branch', 'img','is_active', 'is_inactive','delivery','description','created_at',)
    #Aqui a busca é feito através do campo estrangeiro, primeiro o campo do Model__ depois o campo que quero buscar no outro Model
    search_fields= ['name__name','laptop__title',]
    
    list_filter = ('name','category',)
    
    #importando para excel
    def export_controles_to_excel(request,self,queryset):
    # Cria o workbook e a planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Controle"

        # Adiciona o cabeçalho
        headers = ['Nome', 'Computador',' Departamento' ,'Telefone','Ativo','Filial', 'Data da Entrega', 'Horario da Entrega','Description',]
        worksheet.append(headers)

        # Recupera os dados do modelo e preenche a planilha
        controles = Controle.objects.all() #Busca em Controle todos o objetos
        data_atual = date.today()
        for controle in controles: #Percorre os objetos

           
                
            worksheet.append([
                # controle.id,
                str (controle.name),
                str (controle.laptop),
                str (controle.category),
                str (controle.phones),
                str(controle.is_active),
                str (controle.branch),
              controle.delivery.strftime("%d/%m/%Y"),
              controle.delivery.strftime("%H:%M:%S"),
              str(controle.description)
            ])

        # Configura a resposta HTTP para o download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename=Controles-{data_atual}.xlsx"
        workbook.save(response)
        return response
    
    export_controles_to_excel.short_description = 'Exportar para excel'
    actions = [export_controles_to_excel]
    
#Celulares
class ProductResource(resources.ModelResource):
    class Meta:
        model = Phone

@admin.register(Phone)
class PhoneAdmin(ImportExportModelAdmin):
    list_display = ('title','number','category','brand','imei',)
    search_fields = ('title','number',)
    list_filter = ('is_active', 'brand', 'category','number')

    def export_phone_to_excel(request,self,queryset):
    # Cria o workbook e a planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Controle"

        # Adiciona o cabeçalho
        headers = ['Nome','numero', 'Marca','Armazenamento','Imei','Ativo/Inativo', 'Filial', 'Data da Entrega', 'Horario da Entrega',]
        worksheet.append(headers)

        # Recupera os dados do modelo e preenche a planilha
        phones = Phone.objects.all() #Busca em Controle todos o objetos
        for phone in phones: #Percorre os objetos
  
            #Transforma a palavra True em Ativo, false/inativo
            if phone.is_active == True:
                isActive = 'Ativo'
            else:
                isActive ='Inativo'
                
            #formatted_active = f"{str(phone.is_active)['Ativo']}"
            formatted_phone = f"({str(phone.number)[:2]}) {str(phone.number)[2:7]}-{str(phone.number)[7:]}" # Formata como (xx) xxxxx-xxxx

            worksheet.append([
                # controle.id,
                str (phone.title),
                formatted_phone,
                str(phone.brand),
                str (phone.storage),
                str (phone.imei),
                isActive,
                #formatted_active,
            ])

        # Configura a resposta HTTP para o download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Celular.xlsx"
        workbook.save(response)
        return response
    
    export_phone_to_excel.short_description = 'Exportar para excel'
    actions = [export_phone_to_excel]
    
    # def export_to_csv(self, request, queryset):
    #     response = HttpResponse(content_type='text/csv')
    #     response['Content-Disposition'] = 'attachment; filename="products.csv"'

    #     writer = csv.writer(response)
    #     writer.writerow(['nome', 'marca', 'categoria', 'preço',
    #                      'ativo', 'descrição', 'criado em', 'atualizado em'])
    #     for product in queryset:
    #         writer.writerow([product.title, product.brand.name, product.imei.name,
    #                          product.price, product.is_active, product.description,
    #                          product.created_at, product.updated_at])
    #     return response

    # export_to_csv.short_description = 'Exportar para CSV'
    # actions = [export_to_csv]

#Perifericos
@admin.register(Perifericos)
class PerifeicosAdmin(admin.ModelAdmin):

    list_display = ('title','modelo','amount','brand','is_new','is_used',)
    search_fields = ('title',)
    
     #importando para excel
    def export_periferico_to_excel(request,self,queryset):
    # Cria o workbook e a planilha
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Perifericos"

        # Adiciona o cabeçalho
        headers = ['Nome', 'Modelo', 'Quantidade', 'Marca','Novo', ]
        worksheet.append(headers)

        # Recupera os dados do modelo e preenche a planilha
        perifericos = Perifericos.objects.all() #Busca em Controle todos o objetos
        for periferico in perifericos: #Percorre os objetos
            worksheet.append([
                str (periferico.title),
                str (periferico.modelo),
                str (periferico.amount),
                str (periferico.brand),
            #   controle.delivery.strftime("%Y-%m-%d"),
            #   controle.delivery.strftime("%H:%M:%S"),
            ])

        # Configura a resposta HTTP para o download
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Perifericos.xlsx"
        workbook.save(response)
        return response
    
    export_periferico_to_excel.short_description = 'Exportar para excel'
    actions = [export_periferico_to_excel]
